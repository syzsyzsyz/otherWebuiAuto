#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author : 北京-瑛智
# @wechat : shiyingzhisyz

import json
import logging
import yaml
import toml
import csv
import openpyxl

from configparser import ConfigParser
class DataUtils:
    def __init__(self,file_path,encoding="utf-8",type="json",split="-",sep=",",with_header=True,sheet_name="Sheet1"):
        '''

        :param file_path: 文件路径
        :param encoding: 编码格式
        :param type: 文件类型
        '''
        self.file_path = file_path
        self.encoding = encoding
        self.type = type
        self.split = split
        self.sep = sep
        if type == "tsv":
            self.sep = "\t"
        self.with_header = with_header
        self.sheet_name = sheet_name
    def read_txt(self):
        logging.debug("读取文本文件，%s"%self.file_path)
        with open(self.file_path,"r",encoding=self.encoding) as f:
            return f.read()
    def load_txt(self):
        logging.debug("加载TXT文件，%s"%self.file_path)
        with open(self.file_path,"r",encoding=self.encoding) as f:
            return  [row.strip().split(self.split) for row in f.readlines()]

    def load_yamlOrJsonOrToml(self):
        try:
            logging.debug("读取%s数据文件，%s"%(self.type,self.file_path))
            with open(self.file_path,"r",encoding=self.encoding) as f:
                if self.type == "yaml" :
                    return yaml.full_load(f)
                if self.type == "json":
                    return json.load(f)
                if self.type == "toml":
                    return  toml.load(f)

        except Exception as msg:
            print("异常信息->{0}".format(msg))
        finally:
            f.close()
    def load_csv(self):
        logging.debug("加载%s数据文件,%s"%(self.type,self.file_path))
        data = []
        try:
            with open(self.file_path,encoding=self.encoding) as f:
                if self.with_header:
                    reader = csv.DictReader(f,delimiter=self.sep)
                else:
                    reader = csv.reader(f,delimiter=self.sep)
                for row in reader:
                    data.append(row)
                return  data
        except Exception as msg:
            print("异常信息->{0}".format(msg))
        finally:
            f.close()

    def load_excel(self):
        logging.debug("加载%s数据文件,%s"%(self.type,self.file_path))

        try:
            wb = openpyxl.load_workbook(self.file_path).active

            if self.with_header:
                for row in wb.iter_rows(max_row=1,values_only=True):
                    header_row = row
            data = []
            for row in wb.iter_rows(min_row=2,values_only=True):
                if self.with_header:
                    row_data = dict(zip(header_row,row))
                else:
                    row_data =row
                data.append(row_data)
            return data
        except Exception as msg:
            print("异常信息->{0}".format(msg))
    def load_ini(self):
        logging.debug("加载%s数据文件,%s" % (self.type, self.file_path))
        try:
            conf = ConfigParser()
            conf.read(self.file_path,self.encoding)
            return {section:dict(conf.items(section)) for section in conf.sections()}
        except Exception as msg:
            print("异常信息->{0}".format(msg))


    def load(self):
        file_end_name = str(self.file_path)
        if file_end_name.endswith(".json") or file_end_name.endswith(".yaml") or file_end_name.endswith(".toml"):
            return self.load_yamlOrJsonOrToml()
        if file_end_name.endswith(".txt"):
            return self.load_txt()
        if self.type== "csv" or self.type=="tsv":
            return self.load_csv()
        if self.type=="xlsx":
            return self.load_excel()
        if file_end_name.endswith(".ini"):
            return self.load_ini()
# for i in ["json","yaml","toml","txt","csv","tsv","xlsx","ini"]:
#     data = DataUtils(file_path=DATA_DIR / "test/data.{0}".format(i),encoding="utf-8",type=i,split="-")
#     print(data.load())